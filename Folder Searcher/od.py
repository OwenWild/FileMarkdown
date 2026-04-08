#!/usr/bin/env python3
"""
Local recursive folder analyst.

What this version does differently:
- Scans the full tree recursively for exact counts/sizes/dates.
- Creates one Markdown report per top-level folder inside the root.
- Stores all Markdown files in one output directory.
- Creates one final root report that summarizes those per-folder reports.
- Uses much stronger folder-level reasoning instead of noisy per-file chatter.
- Samples a much larger share of *eligible* files per folder (default 65%), with no hard per-folder cap unless you set one,
  while still capping total evidence so giant folders stay practical.
- Suppresses noisy/broken text extraction and strips replacement characters.
- Avoids over-summarizing low-signal junk like node_modules, caches, venvs,
  build outputs, and similar generated directories.

Default local AI backend: Ollama on http://localhost:11434
"""

from __future__ import annotations

import argparse
import base64
import collections
import contextlib
import html
import json
import math
import mimetypes
import os
import re
import sys
import textwrap
import unicodedata
import warnings
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

try:
    import requests
except ImportError:  # pragma: no cover
    requests = None

try:
    import fitz  # PyMuPDF
except ImportError:  # pragma: no cover
    fitz = None

try:
    from docx import Document as DocxDocument
except ImportError:  # pragma: no cover
    DocxDocument = None

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

try:
    from PIL import Image
except ImportError:  # pragma: no cover
    Image = None


TEXT_EXTENSIONS = {
    ".txt", ".md", ".markdown", ".rst", ".log", ".ini", ".cfg", ".conf", ".toml",
    ".yaml", ".yml", ".json", ".jsonl", ".csv", ".tsv", ".xml", ".html", ".htm",
    ".css", ".js", ".ts", ".jsx", ".tsx", ".py", ".java", ".c", ".cpp", ".h", ".hpp",
    ".cs", ".go", ".rs", ".rb", ".php", ".sql", ".sh", ".bat", ".ps1", ".tex", ".rtf",
}

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp", ".tif", ".tiff", ".heic"}
VIDEO_EXTENSIONS = {".mp4", ".mkv", ".mov", ".avi", ".wmv", ".m4v", ".webm", ".flv", ".mpeg", ".mpg"}
AUDIO_EXTENSIONS = {".mp3", ".wav", ".flac", ".aac", ".m4a", ".ogg", ".wma", ".aiff"}
ARCHIVE_EXTENSIONS = {".zip", ".rar", ".7z", ".tar", ".gz", ".bz2", ".xz", ".iso", ".img"}
DOCUMENT_EXTENSIONS = {".pdf", ".docx", ".doc", ".pptx", ".ppt", ".xlsx", ".xls", ".epub"}
FONT_EXTENSIONS = {".ttf", ".otf", ".woff", ".woff2"}
EXECUTABLE_EXTENSIONS = {".exe", ".msi", ".app", ".dmg", ".pkg", ".apk", ".ipa", ".jar"}
CAD_EXTENSIONS = {".blend", ".obj", ".fbx", ".stl", ".step", ".stp", ".dwg", ".dxf"}

NOISE_DIR_NAMES = {
    "node_modules", "__pycache__", ".git", ".github", ".idea", ".vscode", "dist", "build",
    ".next", ".nuxt", ".cache", "cache", ".venv", "venv", "env", ".mypy_cache",
    ".pytest_cache", "site-packages", "target", "out", "obj", "bin", "Debug", "Release",
    "vendor", "third_party", "deps", "packages", "PackageCache", "coverage", ".parcel-cache",
}

NOISE_FILE_NAMES = {
    "package-lock.json", "yarn.lock", "pnpm-lock.yaml", "poetry.lock", ".ds_store",
}

HIGH_SIGNAL_FILENAMES = {
    "readme.md", "readme.txt", "readme", "package.json", "requirements.txt", "pyproject.toml",
    "cargo.toml", "cmakelists.txt", "dockerfile", "docker-compose.yml", "docker-compose.yaml",
    "vite.config.ts", "vite.config.js", "tsconfig.json", "setup.py", "setup.cfg", "makefile",
    "pom.xml", "build.gradle", "gradle.properties", "go.mod", "go.sum", "composer.json",
    "manifest.json", "config.json", "settings.json",
}

HIGH_SIGNAL_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".txt", ".md", ".json", ".toml", ".yaml", ".yml", ".csv"}
INCOMPLETE_EXTENSIONS = {".crdownload", ".part", ".partial", ".download", ".tmp", ".temp", ".incomplete", ".smdownload"}
ERRORISH_NAME_PATTERNS = [r"_error\.txt$", r"^__.*_error\.txt$", r"^___all_errors\.txt$"]

BADGE_COLORS = {
    "root": "#2563eb",
    "files": "#0f766e",
    "size": "#7c3aed",
    "docs": "#b45309",
    "images": "#be185d",
    "video": "#b91c1c",
    "audio": "#15803d",
    "code": "#374151",
    "updated": "#0369a1",
    "purpose": "#4f46e5",
    "warning": "#b91c1c",
    "note": "#6b7280",
}


def human_size(num_bytes: int) -> str:
    if num_bytes < 1024:
        return f"{num_bytes} B"
    value = float(num_bytes)
    for unit in ["KB", "MB", "GB", "TB", "PB"]:
        value /= 1024.0
        if value < 1024 or unit == "PB":
            return f"{value:.2f} {unit}"
    return f"{num_bytes} B"


def format_dt(ts: Optional[float]) -> Optional[str]:
    if ts is None:
        return None
    try:
        return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return None


def safe_creation_timestamp(stat_result: os.stat_result) -> Optional[float]:
    if hasattr(stat_result, "st_birthtime"):
        return getattr(stat_result, "st_birthtime")
    return getattr(stat_result, "st_ctime", None)


def truncate(text: str, max_chars: int) -> str:
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) <= max_chars:
        return text
    return text[: max_chars - 1].rstrip() + "…"


def clean_text(text: str, max_chars: int = 12000) -> str:
    if not text:
        return ""
    text = unicodedata.normalize("NFKC", text)
    text = text.replace("\x00", " ").replace("\ufffd", " ")
    text = "".join(ch if (ch.isprintable() or ch in "\n\r\t") else " " for ch in text)
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"\s([,.;:!?])", r"\1", text)
    return truncate(text, max_chars)


def likely_binary_bytes(raw: bytes) -> bool:
    if not raw:
        return False
    nul_ratio = raw.count(b"\x00") / max(1, len(raw))
    if nul_ratio > 0.02:
        return True
    printable = 0
    for b in raw[:4000]:
        if 32 <= b <= 126 or b in (9, 10, 13):
            printable += 1
    ratio = printable / max(1, min(len(raw), 4000))
    return ratio < 0.55


def markdown_escape(text: str) -> str:
    return text.replace("|", "\\|")


def badge(text: str, color: str) -> str:
    escaped = html.escape(text)
    return (
        f"<span style=\"display:inline-block;padding:0.22em 0.58em;margin:0.1em 0.18em 0.1em 0;"
        f"background:{color};color:#fff;border-radius:999px;font-size:0.9em;font-weight:600;\">{escaped}</span>"
    )


def is_hidden_path(path: Path) -> bool:
    return any(part.startswith(".") for part in path.parts if part not in (".", ".."))


def guess_category(path: Path, mime_type: str) -> str:
    ext = path.suffix.lower()
    if ext in IMAGE_EXTENSIONS or mime_type.startswith("image/"):
        return "image"
    if ext in VIDEO_EXTENSIONS or mime_type.startswith("video/"):
        return "video"
    if ext in AUDIO_EXTENSIONS or mime_type.startswith("audio/"):
        return "audio"
    if ext in ARCHIVE_EXTENSIONS:
        return "archive"
    if ext in DOCUMENT_EXTENSIONS:
        return "document"
    if ext in FONT_EXTENSIONS:
        return "font"
    if ext in EXECUTABLE_EXTENSIONS or mime_type in {"application/x-msdownload", "application/x-dosexec"}:
        return "executable"
    if ext in CAD_EXTENSIONS:
        return "3d/cad"
    if ext in TEXT_EXTENSIONS or mime_type.startswith("text/"):
        if ext in {".py", ".js", ".ts", ".jsx", ".tsx", ".java", ".c", ".cpp", ".h", ".hpp", ".cs", ".go", ".rs", ".rb", ".php", ".sql", ".sh", ".bat", ".ps1", ".css", ".html", ".htm", ".xml", ".json", ".yaml", ".yml", ".toml"}:
            return "code/config"
        return "text/data"
    return "other"


def sample_indices(total: int, wanted: int) -> List[int]:
    if total <= 0:
        return []
    if total <= wanted:
        return list(range(total))
    if wanted == 1:
        return [0]
    step = (total - 1) / (wanted - 1)
    values = sorted({round(i * step) for i in range(wanted)})
    return [min(total - 1, max(0, i)) for i in values]


def safe_slug(text: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9._-]+", "-", text.strip())
    slug = re.sub(r"-+", "-", slug).strip("-._")
    return slug or "Folder"


def safe_read_text_file(path: Path, max_chars: int = 12000) -> str:
    raw = path.read_bytes()[: max_chars * 6]
    if likely_binary_bytes(raw):
        return ""
    for encoding in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
        try:
            return clean_text(raw.decode(encoding), max_chars=max_chars)
        except Exception:
            continue
    return ""


def summarize_tabular_preview(rows: List[List[Any]], max_rows: int = 10, max_cols: int = 8) -> str:
    lines: List[str] = []
    for row in rows[:max_rows]:
        clean = [truncate(str(cell), 60) for cell in row[:max_cols] if cell not in (None, "")]
        if clean:
            lines.append(" | ".join(clean))
    return clean_text(" ; ".join(lines), max_chars=2500)


@dataclass
class FileRecord:
    rel_path: str
    abs_path: str
    name: str
    extension: str
    category: str
    size_bytes: int
    created: Optional[str]
    modified: Optional[str]
    mime_type: str
    is_hidden: bool
    top_folder: str
    is_noise_path: bool = False
    excerpt: Optional[str] = None
    image_description: Optional[str] = None
    width: Optional[int] = None
    height: Optional[int] = None
    parse_error: Optional[str] = None


@dataclass
class FolderStats:
    rel_path: str
    depth: int
    total_size_bytes: int = 0
    file_count: int = 0
    category_counts: collections.Counter = field(default_factory=collections.Counter)
    extension_counts: collections.Counter = field(default_factory=collections.Counter)
    latest_modified: Optional[str] = None
    earliest_created: Optional[str] = None
    largest_files: List[Tuple[str, int]] = field(default_factory=list)
    recent_files: List[Tuple[str, str]] = field(default_factory=list)


class LocalAIError(RuntimeError):
    pass


def strip_model_wrappers(text: str) -> str:
    if not text:
        return ""
    text = text.strip().replace("﻿", "")
    text = re.sub(r"<think>.*?</think>", " ", text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r"```(?:json)?", "", text, flags=re.IGNORECASE)
    return text.strip()


def extract_json_text(text: str) -> str:
    cleaned = strip_model_wrappers(text)
    if not cleaned:
        return ""
    stripped = cleaned.strip()
    if stripped[:1] in "[{" and stripped[-1:] in "]}":
        return stripped
    decoder = json.JSONDecoder()
    for idx, ch in enumerate(stripped):
        if ch not in "[{":
            continue
        try:
            _, end = decoder.raw_decode(stripped[idx:])
            return stripped[idx: idx + end]
        except Exception:
            continue
    return stripped


def coerce_json_result(text: str) -> Any:
    candidate = extract_json_text(text)
    if not candidate:
        raise json.JSONDecodeError("empty", text, 0)
    return json.loads(candidate)


def compact_for_ai(value: Any, max_items: int = 10, max_str: int = 240) -> Any:
    if isinstance(value, dict):
        return {str(key): compact_for_ai(item, max_items=max_items, max_str=max_str) for key, item in value.items()}
    if isinstance(value, list):
        return [compact_for_ai(item, max_items=max_items, max_str=max_str) for item in value[:max_items]]
    if isinstance(value, str):
        return truncate(clean_text(value, max_chars=max_str), max_str)
    return value


class OllamaClient:
    def __init__(self, base_url: str, model: str, timeout: int = 240, keep_alive: str = "30m", num_ctx: int = 32768) -> None:
        if requests is None:
            raise LocalAIError("Install 'requests' to use local AI: pip install requests")
        self.base_url = base_url.rstrip("/")
        self.model = model
        self.timeout = timeout
        self.keep_alive = keep_alive
        self.num_ctx = num_ctx

    def _post_generate(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        try:
            response = requests.post(f"{self.base_url}/api/generate", json=payload, timeout=self.timeout)
            response.raise_for_status()
        except Exception as exc:
            raise LocalAIError(f"Could not reach Ollama at {self.base_url}: {exc}") from exc
        data = response.json()
        if isinstance(data, dict) and data.get("error"):
            raise LocalAIError(str(data["error"]))
        return data

    def generate_text(
        self,
        prompt: str,
        system: Optional[str] = None,
        images: Optional[List[Path]] = None,
        schema: Optional[Dict[str, Any]] = None,
        temperature: float = 0.1,
    ) -> str:
        payload: Dict[str, Any] = {
            "model": self.model,
            "prompt": prompt,
            "stream": False,
            "keep_alive": self.keep_alive,
            "options": {
                "temperature": temperature,
                "num_ctx": self.num_ctx,
                "num_predict": 1024,
            },
        }
        if system:
            payload["system"] = system
        if schema:
            payload["format"] = schema
        if images:
            payload["images"] = [base64.b64encode(p.read_bytes()).decode("utf-8") for p in images]

        attempts: List[Dict[str, Any]] = [payload]
        if schema:
            retry_payload = dict(payload)
            retry_payload.pop("format", None)
            retry_payload["prompt"] = (
                f"{prompt}\n\nReturn only valid JSON. No markdown fences. No commentary before or after the JSON."
            )
            attempts.append(retry_payload)

        last_raw = ""
        for attempt in attempts:
            data = self._post_generate(attempt)
            text = data.get("response") or data.get("message", {}).get("content") or ""
            cleaned = strip_model_wrappers(text)
            if cleaned:
                return cleaned
            last_raw = text or json.dumps(data)[:400]
        raise LocalAIError(f"The local model returned an empty response. Raw preview: {last_raw[:300]!r}")

    def generate_json(
        self,
        prompt: str,
        system: Optional[str] = None,
        images: Optional[List[Path]] = None,
        schema: Optional[Dict[str, Any]] = None,
        temperature: float = 0.1,
    ) -> Any:
        errors: List[str] = []
        for use_schema in ([True, False] if schema else [False]):
            try:
                text = self.generate_text(
                    prompt=prompt,
                    system=system,
                    images=images,
                    schema=schema if use_schema else None,
                    temperature=temperature,
                )
                return coerce_json_result(text)
            except Exception as exc:
                errors.append(str(exc))
                continue
        raise LocalAIError("Model did not return valid JSON. " + " | ".join(errors[:2]))


def add_largest_file(bucket: List[Tuple[str, int]], rel_path: str, size_bytes: int, limit: int = 12) -> None:
    bucket.append((rel_path, size_bytes))
    bucket.sort(key=lambda item: item[1], reverse=True)
    del bucket[limit:]


def add_recent_file(bucket: List[Tuple[str, str]], rel_path: str, modified: Optional[str], limit: int = 12) -> None:
    if not modified:
        return
    bucket.append((rel_path, modified))
    bucket.sort(key=lambda item: item[1], reverse=True)
    del bucket[limit:]


def path_is_noise(rel_path: Path) -> bool:
    lowered = [part.lower() for part in rel_path.parts]
    return any(part in {name.lower() for name in NOISE_DIR_NAMES} for part in lowered[:-1])


def file_name_is_errorish(name: str) -> bool:
    lowered = name.lower()
    return any(re.search(pattern, lowered) for pattern in ERRORISH_NAME_PATTERNS)


def is_high_signal_file(path: Path) -> bool:
    name = path.name.lower()
    if name in HIGH_SIGNAL_FILENAMES:
        return True
    if path.suffix.lower() in HIGH_SIGNAL_EXTENSIONS and not path_is_noise(path):
        return True
    if name.endswith(".sln") or name.endswith(".uproject") or name.endswith(".ipynb"):
        return True
    return False


def should_extract_excerpt(path: Path, category: str, size_bytes: int, max_text_file_mb: int) -> bool:
    if size_bytes > max_text_file_mb * 1024 * 1024:
        return False
    name = path.name.lower()
    if name in NOISE_FILE_NAMES:
        return False
    if category == "document":
        return True
    if category == "text/data":
        return True
    if category == "code/config" and is_high_signal_file(path):
        return True
    return False


def extract_pdf_excerpt(path: Path, max_chars: int = 12000, max_pages_sampled: int = 12) -> str:
    if fitz is None:
        return ""
    try:
        doc = fitz.open(path)
    except Exception:
        return ""
    text_parts: List[str] = []
    try:
        page_numbers = sample_indices(len(doc), max_pages_sampled)
        for idx in page_numbers:
            try:
                page = doc[idx]
                text = page.get_text("text", sort=True)
                if text:
                    text_parts.append(text)
            except Exception:
                continue
    finally:
        doc.close()
    return clean_text("\n\n".join(text_parts), max_chars=max_chars)


def extract_docx_excerpt(path: Path, max_chars: int = 12000) -> str:
    if DocxDocument is None:
        return ""
    try:
        doc = DocxDocument(str(path))
        text = "\n".join(p.text for p in doc.paragraphs if p.text and p.text.strip())
        return clean_text(text, max_chars=max_chars)
    except Exception:
        return ""


def extract_xlsx_excerpt(path: Path, max_chars: int = 12000) -> str:
    if load_workbook is None:
        return ""
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message=r"Print area cannot be set to Defined name: .*",
                category=UserWarning,
            )
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception:
        return ""

    fragments: List[str] = []
    try:
        for ws in list(wb.worksheets)[:4]:
            rows: List[List[Any]] = []
            for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
                rows.append(list(row[:8]))
            preview = summarize_tabular_preview(rows)
            if preview:
                fragments.append(f"Sheet '{ws.title}': {preview}")
    finally:
        wb.close()
    return clean_text("\n".join(fragments), max_chars=max_chars)


def extract_textual_excerpt(path: Path, max_chars: int = 12000) -> str:
    ext = path.suffix.lower()
    if ext == ".pdf":
        return extract_pdf_excerpt(path, max_chars=max_chars)
    if ext == ".docx":
        return extract_docx_excerpt(path, max_chars=max_chars)
    if ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return extract_xlsx_excerpt(path, max_chars=max_chars)
    if ext in TEXT_EXTENSIONS:
        return safe_read_text_file(path, max_chars=max_chars)
    if fitz is not None and ext in {".epub", ".xps", ".oxps", ".cbz", ".fb2"}:
        return extract_pdf_excerpt(path, max_chars=max_chars)
    return ""


def get_image_dimensions(path: Path) -> Tuple[Optional[int], Optional[int]]:
    if Image is None:
        return None, None
    try:
        with Image.open(path) as img:
            return img.width, img.height
    except Exception:
        return None, None


class FolderScanner:
    def __init__(self, args: argparse.Namespace) -> None:
        self.args = args
        self.root = Path(args.root).resolve()
        self.root_name = self.root.name or str(self.root)
        self.folder_stats: Dict[str, FolderStats] = {}
        self.file_records: List[FileRecord] = []
        self.top_folder_files: Dict[str, List[FileRecord]] = collections.defaultdict(list)
        self.top_level_folders: List[str] = []
        self.output_dir_resolved = Path(args.output_dir).resolve()

    def ensure_folder(self, rel_path: str) -> FolderStats:
        if rel_path not in self.folder_stats:
            depth = 0 if rel_path == "." else rel_path.count(os.sep) + 1
            self.folder_stats[rel_path] = FolderStats(rel_path=rel_path, depth=depth)
        return self.folder_stats[rel_path]

    def should_skip_for_output_loop(self, path: Path) -> bool:
        try:
            if self.args.exclude_output_dir_from_scan and self.output_dir_resolved in path.parents:
                return True
        except Exception:
            pass
        return False

    def iter_paths(self) -> Iterable[Path]:
        for path in self.root.rglob("*"):
            rel = path.relative_to(self.root)
            if self.should_skip_for_output_loop(path):
                continue
            if not self.args.include_hidden and is_hidden_path(rel):
                continue
            if path.is_file():
                yield path

    def ancestor_keys(self, rel_path: Path) -> List[str]:
        keys = ["."]
        for parent in reversed(rel_path.parents):
            if str(parent) in {"", "."}:
                continue
            keys.append(str(parent))
        return keys

    def scan(self) -> Dict[str, Any]:
        self.ensure_folder(".")
        total_files = 0
        total_dirs = 0
        top_level_folders: List[str] = []

        for dirpath, dirnames, _ in os.walk(self.root):
            dir_path = Path(dirpath)
            if self.should_skip_for_output_loop(dir_path):
                dirnames[:] = []
                continue
            rel_dir = os.path.relpath(dirpath, self.root)
            if rel_dir == ".":
                rel_dir = "."
            if not self.args.include_hidden:
                dirnames[:] = [d for d in dirnames if not d.startswith(".")]
            if rel_dir != ".":
                total_dirs += 1
                self.ensure_folder(rel_dir)
                if os.sep not in rel_dir:
                    top_level_folders.append(rel_dir)

        self.top_level_folders = sorted(set(top_level_folders), key=lambda x: x.lower())

        for path in self.iter_paths():
            total_files += 1
            rel_path = path.relative_to(self.root)
            rel_str = str(rel_path)
            mime_type = mimetypes.guess_type(str(path))[0] or "application/octet-stream"
            category = guess_category(path, mime_type)
            hidden = is_hidden_path(rel_path)

            try:
                st = path.stat()
            except OSError:
                continue

            top_folder = rel_path.parts[0] if len(rel_path.parts) > 1 else "."
            created = format_dt(safe_creation_timestamp(st))
            modified = format_dt(st.st_mtime)
            size_bytes = st.st_size
            noise = path_is_noise(rel_path)

            record = FileRecord(
                rel_path=rel_str,
                abs_path=str(path),
                name=path.name,
                extension=path.suffix.lower() or "(none)",
                category=category,
                size_bytes=size_bytes,
                created=created,
                modified=modified,
                mime_type=mime_type,
                is_hidden=hidden,
                top_folder=top_folder,
                is_noise_path=noise,
            )
            if category == "image":
                record.width, record.height = get_image_dimensions(path)

            if should_extract_excerpt(path, category, size_bytes, self.args.max_text_file_mb) and not noise:
                try:
                    excerpt = extract_textual_excerpt(path, max_chars=self.args.max_excerpt_chars)
                    if excerpt:
                        record.excerpt = excerpt
                except Exception as exc:
                    record.parse_error = f"Extraction failed: {exc}"

            self.file_records.append(record)
            self.top_folder_files[top_folder].append(record)

            for key in self.ancestor_keys(rel_path):
                fs = self.ensure_folder(key)
                fs.file_count += 1
                fs.total_size_bytes += size_bytes
                fs.category_counts[category] += 1
                fs.extension_counts[record.extension] += 1
                add_largest_file(fs.largest_files, rel_str, size_bytes, limit=self.args.max_largest_files)
                add_recent_file(fs.recent_files, rel_str, modified, limit=self.args.max_recent_files)
                if modified and (fs.latest_modified is None or modified > fs.latest_modified):
                    fs.latest_modified = modified
                if created and (fs.earliest_created is None or created < fs.earliest_created):
                    fs.earliest_created = created

        root_stats = self.folder_stats["."]
        return {
            "root_path": str(self.root),
            "root_name": self.root_name,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "total_files": total_files,
            "total_dirs": total_dirs,
            "total_size_bytes": root_stats.total_size_bytes,
            "top_level_folder_count": len(self.top_level_folders),
        }


def category_size_breakdown(records: Sequence[FileRecord]) -> List[Tuple[str, int, int]]:
    counts: Dict[str, Tuple[int, int]] = {}
    for rec in records:
        count, size = counts.get(rec.category, (0, 0))
        counts[rec.category] = (count + 1, size + rec.size_bytes)
    rows = [(cat, count, size) for cat, (count, size) in counts.items()]
    rows.sort(key=lambda row: row[2], reverse=True)
    return rows


def top_extensions(records: Sequence[FileRecord], limit: int = 10) -> List[Tuple[str, int]]:
    counter = collections.Counter(rec.extension for rec in records)
    return counter.most_common(limit)


def score_record_for_evidence(rec: FileRecord) -> Tuple[int, int, int, str]:
    path = Path(rec.rel_path)
    score = 0
    if is_high_signal_file(path):
        score += 120
    if rec.category == "document":
        score += 90
    elif rec.category == "text/data":
        score += 70
    elif rec.category == "archive":
        score += 60
    elif rec.category == "executable":
        score += 55
    elif rec.category == "image":
        score += 35
    elif rec.category == "audio":
        score += 25
    elif rec.category == "video":
        score += 25
    elif rec.category == "code/config" and is_high_signal_file(path):
        score += 75
    if rec.is_noise_path:
        score -= 80
    if file_name_is_errorish(rec.name):
        score += 20
    if rec.extension in INCOMPLETE_EXTENSIONS:
        score += 35
    return (score, rec.size_bytes, 1 if rec.excerpt else 0, rec.rel_path.lower())


def select_folder_evidence(records: Sequence[FileRecord], ratio: float, max_items: int) -> List[FileRecord]:
    eligible = [r for r in records if not r.is_noise_path and r.name.lower() not in NOISE_FILE_NAMES]
    if not eligible:
        eligible = list(records)
    if not eligible:
        return []

    ratio = max(0.01, min(1.0, float(ratio)))
    target = max(1, math.ceil(len(eligible) * ratio))
    if max_items and max_items > 0:
        target = min(target, max_items)

    important = [r for r in eligible if is_high_signal_file(Path(r.rel_path))]
    important_sorted = sorted(important, key=score_record_for_evidence, reverse=True)
    selected: List[FileRecord] = important_sorted[:target]
    selected_set = {r.rel_path for r in selected}

    remaining = [r for r in sorted(eligible, key=score_record_for_evidence, reverse=True) if r.rel_path not in selected_set]
    needed = max(0, target - len(selected))
    if needed and remaining:
        for idx in sample_indices(len(remaining), needed):
            selected.append(remaining[idx])

    selected.sort(key=lambda r: score_record_for_evidence(r), reverse=True)
    return selected[:target]


def direct_child_folder_stats(folder_stats: Dict[str, FolderStats], parent_rel: str, limit: int = 10) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    parent_depth = 0 if parent_rel == "." else parent_rel.count(os.sep) + 1
    for rel, fs in folder_stats.items():
        if rel == ".":
            continue
        if parent_rel == ".":
            if os.sep not in rel:
                rows.append({
                    "path": rel,
                    "files": fs.file_count,
                    "size_bytes": fs.total_size_bytes,
                    "size_human": human_size(fs.total_size_bytes),
                })
        else:
            if Path(rel).parent.as_posix().replace("/", os.sep) == parent_rel and fs.depth == parent_depth + 1:
                rows.append({
                    "path": rel,
                    "files": fs.file_count,
                    "size_bytes": fs.total_size_bytes,
                    "size_human": human_size(fs.total_size_bytes),
                })
    rows.sort(key=lambda item: item["size_bytes"], reverse=True)
    return rows[:limit]


def iso_to_age_bucket(dt_str: Optional[str]) -> str:
    if not dt_str:
        return "unknown"
    try:
        dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
        age_days = (datetime.now(timezone.utc) - dt).days
    except Exception:
        return "unknown"
    if age_days <= 7:
        return "last 7 days"
    if age_days <= 30:
        return "last 30 days"
    if age_days <= 180:
        return "last 6 months"
    if age_days <= 365:
        return "last year"
    return "older"


def maybe_describe_images(ai: Optional[OllamaClient], records: Sequence[FileRecord], max_images: int) -> List[Dict[str, str]]:
    chosen = [r for r in records if r.category == "image"][:max_images]
    if not chosen:
        return []
    results: List[Dict[str, str]] = []
    if ai is None:
        for r in chosen:
            desc = []
            if r.width and r.height:
                desc.append(f"{r.width}×{r.height}")
            desc.append(r.extension)
            results.append({"path": r.rel_path, "summary": ", ".join(desc)})
        return results

    system = (
        "You describe images for a local folder inventory. "
        "Be concise, literal, and non-speculative. "
        "Do not identify people. Do not mention style unless obvious."
    )
    schema = {
        "type": "object",
        "properties": {"description": {"type": "string"}},
        "required": ["description"],
    }
    for r in chosen:
        try:
            data = ai.generate_json(
                prompt=textwrap.dedent(
                    f"""
                    Describe this image for a folder inventory.
                    File: {r.rel_path}
                    Return JSON with one short factual sentence in 'description'.
                    """
                ).strip(),
                system=system,
                images=[Path(r.abs_path)],
                schema=schema,
                temperature=0.05,
            )
            desc = clean_text(str(data.get("description", "")), max_chars=180)
            if desc:
                results.append({"path": r.rel_path, "summary": desc})
        except Exception:
            continue
    return results


def build_folder_evidence(
    folder_rel: str,
    records: Sequence[FileRecord],
    folder_stats: Dict[str, FolderStats],
    args: argparse.Namespace,
    ai: Optional[OllamaClient],
) -> Dict[str, Any]:
    fs = folder_stats[folder_rel]
    selected = select_folder_evidence(records, ratio=args.folder_analysis_ratio, max_items=args.max_evidence_files_per_folder)

    document_notes: List[Dict[str, str]] = []
    for rec in selected:
        if rec.excerpt and len(document_notes) < args.max_doc_evidence_per_folder:
            document_notes.append({
                "path": rec.rel_path,
                "summary": truncate(rec.excerpt, 320),
            })

    image_notes = maybe_describe_images(ai if args.analyze_images else None, selected, args.max_image_evidence_per_folder)

    representative_files = [r.rel_path for r in selected[: args.max_representative_files]]
    manifests = [r.rel_path for r in selected if is_high_signal_file(Path(r.rel_path))][:20]

    cleanup_candidates = []
    for rec in sorted(records, key=lambda r: r.size_bytes, reverse=True):
        lower_name = rec.name.lower()
        if rec.extension in INCOMPLETE_EXTENSIONS:
            cleanup_candidates.append({"path": rec.rel_path, "reason": "incomplete or partial download", "size": human_size(rec.size_bytes)})
        elif file_name_is_errorish(lower_name):
            cleanup_candidates.append({"path": rec.rel_path, "reason": "error log or failed transfer note", "size": human_size(rec.size_bytes)})
        elif rec.category in {"archive", "executable"} and iso_to_age_bucket(rec.modified) in {"older", "last year"}:
            cleanup_candidates.append({"path": rec.rel_path, "reason": "old installer or archive in Downloads", "size": human_size(rec.size_bytes)})
        if len(cleanup_candidates) >= 16:
            break

    evidence = {
        "folder_path": folder_rel,
        "folder_name": Path(folder_rel).name,
        "file_count": fs.file_count,
        "total_size_bytes": fs.total_size_bytes,
        "total_size_human": human_size(fs.total_size_bytes),
        "latest_modified": fs.latest_modified,
        "earliest_created": fs.earliest_created,
        "category_breakdown": [
            {"category": c, "files": n, "size": human_size(sz), "size_bytes": sz}
            for c, n, sz in category_size_breakdown(records)
        ],
        "top_extensions": [{"extension": ext, "count": count} for ext, count in top_extensions(records, 12)],
        "largest_files": [
            {"path": path, "size": human_size(size), "size_bytes": size}
            for path, size in fs.largest_files[: args.max_largest_files]
        ],
        "recent_files": [{"path": path, "modified": modified} for path, modified in fs.recent_files[: args.max_recent_files]],
        "direct_child_folders": direct_child_folder_stats(folder_stats, folder_rel, limit=10),
        "representative_files": representative_files,
        "signal_files": manifests,
        "document_notes": document_notes,
        "image_notes": image_notes,
        "cleanup_candidates": cleanup_candidates,
        "analysis_ratio_requested": args.folder_analysis_ratio,
        "eligible_file_estimate": len([r for r in records if not r.is_noise_path]),
        "selected_evidence_count": len(selected),
    }
    return evidence


def build_folder_ai_payload(evidence: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "folder_path": evidence["folder_path"],
        "folder_name": evidence["folder_name"],
        "file_count": evidence["file_count"],
        "total_size_human": evidence["total_size_human"],
        "latest_modified": evidence["latest_modified"],
        "earliest_created": evidence["earliest_created"],
        "category_breakdown": compact_for_ai(evidence["category_breakdown"], max_items=8, max_str=80),
        "top_extensions": compact_for_ai(evidence["top_extensions"], max_items=10, max_str=40),
        "direct_child_folders": compact_for_ai(evidence["direct_child_folders"], max_items=8, max_str=80),
        "largest_files": compact_for_ai(evidence["largest_files"], max_items=10, max_str=100),
        "representative_files": compact_for_ai(evidence["representative_files"], max_items=60, max_str=120),
        "signal_files": compact_for_ai(evidence["signal_files"], max_items=18, max_str=140),
        "document_notes": compact_for_ai(evidence["document_notes"], max_items=8, max_str=220),
        "image_notes": compact_for_ai(evidence["image_notes"], max_items=6, max_str=180),
        "cleanup_candidates": compact_for_ai(evidence["cleanup_candidates"], max_items=8, max_str=140),
        "selected_evidence_count": evidence["selected_evidence_count"],
        "eligible_file_estimate": evidence["eligible_file_estimate"],
    }


def build_root_ai_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    child_folders = []
    for item in payload.get("child_folders", [])[:18]:
        child_folders.append({
            "folder_path": item.get("folder_path"),
            "label": item.get("label") or item.get("folder_path"),
            "file_count": item.get("file_count"),
            "size_human": item.get("size_human"),
            "summary": truncate(clean_text(item.get("summary", ""), max_chars=260), 260),
            "purpose": truncate(clean_text(item.get("purpose", ""), max_chars=260), 260),
            "top_tags": compact_for_ai(item.get("top_tags", []), max_items=6, max_str=40),
        })
    return {
        "overview": compact_for_ai(payload.get("overview", {}), max_items=20, max_str=140),
        "root_stats": compact_for_ai(payload.get("root_stats", {}), max_items=10, max_str=80),
        "child_folders": child_folders,
        "loose_root_files": compact_for_ai(payload.get("loose_root_files", []), max_items=12, max_str=120),
        "cleanup_candidates": compact_for_ai(payload.get("cleanup_candidates", {}), max_items=10, max_str=140),
    }


def generate_folder_narrative(ai: Optional[OllamaClient], evidence: Dict[str, Any]) -> Dict[str, Any]:
    if ai is None:
        dominant = evidence["category_breakdown"][:2]
        dom_text = ", ".join(f"{d['category']} ({d['files']})" for d in dominant) if dominant else "mixed files"
        return {
            "title": f"{evidence['folder_name']} — Folder Markup",
            "label": evidence["folder_name"],
            "summary": f"A large folder with {evidence['file_count']} files, dominated by {dom_text}.",
            "purpose": "Purpose inferred deterministically because AI is disabled.",
            "what_is_here": "This folder report is based on exact file counts, types, dates, and selected representative files.",
            "top_tags": [d["category"] for d in evidence["category_breakdown"][:5]],
            "keep_delete_move": [],
        }

    system = (
        "You write restrained, accurate folder summaries for a local file organizer. "
        "Your job is to infer what a folder is for using folder names, file names, type mix, direct-child folders, a controlled sample of representative files, and a few document/image notes. "
        "Do not over-focus on one random file. Do not hallucinate. Ignore obvious generated dependency folders unless they are central to the folder's purpose."
    )
    schema = {
        "type": "object",
        "properties": {
            "title": {"type": "string"},
            "label": {"type": "string"},
            "summary": {"type": "string"},
            "purpose": {"type": "string"},
            "what_is_here": {"type": "string"},
            "top_tags": {"type": "array", "items": {"type": "string"}},
            "keep_delete_move": {"type": "array", "items": {"type": "string"}},
        },
        "required": ["title", "label", "summary", "purpose", "what_is_here", "top_tags", "keep_delete_move"],
    }
    ai_payload = build_folder_ai_payload(evidence)
    prompt = textwrap.dedent(
        f"""
        Write a clean Markdown-ready narrative for this folder.

        Rules:
        - Focus on the folder's overall purpose, not a single weird file.
        - Use only the supplied evidence.
        - If this looks like a project, infer its purpose from the folder name, high-signal file names, representative file names, and document notes.
        - If this looks like a media library, installer dump, AI models directory, personal archive, school materials, or mixed downloads folder, say so clearly when supported.
        - 'keep_delete_move' should be short suggestions, not commands, and only when justified.
        - Keep the writing tight. No fluff.

        Folder evidence JSON:
        {json.dumps(ai_payload, ensure_ascii=False)}
        """
    ).strip()
    return ai.generate_json(prompt=prompt, system=system, schema=schema, temperature=0.12)


def build_root_cleanup_candidates(records: Sequence[FileRecord]) -> Dict[str, List[Dict[str, str]]]:
    incomplete: List[Dict[str, str]] = []
    old_installers: List[Dict[str, str]] = []
    error_logs: List[Dict[str, str]] = []
    duplicates_map: Dict[Tuple[str, int], List[FileRecord]] = collections.defaultdict(list)

    for rec in records:
        duplicates_map[(rec.name.lower(), rec.size_bytes)].append(rec)
        if rec.extension in INCOMPLETE_EXTENSIONS and len(incomplete) < 20:
            incomplete.append({"path": rec.rel_path, "size": human_size(rec.size_bytes), "reason": "partial/incomplete download"})
        elif file_name_is_errorish(rec.name) and len(error_logs) < 20:
            error_logs.append({"path": rec.rel_path, "size": human_size(rec.size_bytes), "reason": "error log / failed transfer"})
        elif rec.category in {"archive", "executable"} and iso_to_age_bucket(rec.modified) in {"older", "last year"} and len(old_installers) < 20:
            old_installers.append({"path": rec.rel_path, "size": human_size(rec.size_bytes), "reason": "older installer/archive"})

    duplicates: List[Dict[str, str]] = []
    for (_, size), group in sorted(duplicates_map.items(), key=lambda item: (-len(item[1]), -item[0][1])):
        if len(group) >= 2:
            duplicates.append({
                "name": group[0].name,
                "size": human_size(size),
                "copies": str(len(group)),
                "examples": ", ".join(g.rel_path for g in group[:3]),
            })
        if len(duplicates) >= 15:
            break

    return {
        "incomplete_downloads": incomplete,
        "older_installers_or_archives": old_installers,
        "error_logs": error_logs,
        "possible_duplicates": duplicates,
    }


def generate_root_narrative(ai: Optional[OllamaClient], payload: Dict[str, Any]) -> Dict[str, Any]:
    if ai is None:
        return {
            "title": f"{payload['overview']['root_name']} — Full Folder Markup",
            "summary": "This root summary was built deterministically because AI is disabled.",
            "purpose": "A high-level summary of the root folder and its major subfolders.",
            "top_tags": [(item.get("label") or item.get("folder_path") or "folder") for item in payload.get("child_folders", [])[:5]],
            "organization_summary": "The report groups the root by major top-level folders and highlights large or review-worthy items.",
            "delete_recommendations": ["Review incomplete downloads and stale installers first."],
            "move_recommendations": ["Move stable long-term collections out of Downloads into dedicated library folders."],
            "priority_actions": ["Clear partial downloads.", "Move long-term archives.", "Re-run after cleanup."],
        }

    system = (
        "You write a cohesive top-level cleanup and organization summary for a Downloads folder. "
        "Use exact metrics from the payload and the already-computed child folder narratives. "
        "Be practical and conservative. Prioritize cleanup suggestions that are low-risk: incomplete downloads, stale installers, obvious duplicates, misplaced large archives, and giant long-term libraries that should not live in Downloads."
    )
    schema = {
        "type": "object",
        "properties": {
            "title": {"type": "string"},
            "summary": {"type": "string"},
            "purpose": {"type": "string"},
            "top_tags": {"type": "array", "items": {"type": "string"}},
            "organization_summary": {"type": "string"},
            "delete_recommendations": {"type": "array", "items": {"type": "string"}},
            "move_recommendations": {"type": "array", "items": {"type": "string"}},
            "priority_actions": {"type": "array", "items": {"type": "string"}},
        },
        "required": [
            "title", "summary", "purpose", "top_tags", "organization_summary",
            "delete_recommendations", "move_recommendations", "priority_actions"
        ],
    }
    ai_payload = build_root_ai_payload(payload)
    prompt = textwrap.dedent(
        f"""
        Write a polished root-level Markdown narrative for this folder inventory.

        Rules:
        - Summarize the major top-level folders as a cohesive whole.
        - Make cleanup recommendations only when supported by the payload.
        - Make move recommendations that improve organization, but keep them realistic.
        - Keep the tone practical, not preachy.
        - Avoid inventing exact paths unless they already exist in the data; you can recommend folder categories like 'Music archive' or 'Software installers'.

        Payload JSON:
        {json.dumps(ai_payload, ensure_ascii=False)}
        """
    ).strip()
    return ai.generate_json(prompt=prompt, system=system, schema=schema, temperature=0.14)


def render_folder_markdown(evidence: Dict[str, Any], narrative: Dict[str, Any]) -> str:
    lines: List[str] = []
    lines.append(f"# {narrative['title']}")
    lines.append("")
    lines.append(
        " ".join(
            [
                badge(f"{evidence['file_count']} files", BADGE_COLORS["files"]),
                badge(evidence['total_size_human'], BADGE_COLORS["size"]),
                badge(f"evidence: {evidence['selected_evidence_count']} files", BADGE_COLORS["root"]),
                badge(f"updated {evidence['latest_modified'] or 'unknown'}", BADGE_COLORS["updated"]),
            ]
        )
    )
    lines.append("")
    if narrative.get("top_tags"):
        lines.append(" ".join(badge(tag, BADGE_COLORS["purpose"]) for tag in narrative["top_tags"][:8]))
        lines.append("")
    lines.append(f"> {narrative['summary']}")
    lines.append("")
    lines.append(f"**Likely purpose:** {narrative['purpose']}")
    lines.append("")
    lines.append(narrative["what_is_here"])
    lines.append("")

    lines.append("## Snapshot")
    lines.append("")
    lines.append("| Metric | Value |")
    lines.append("|---|---:|")
    lines.append(f"| Folder path | `{markdown_escape(evidence['folder_path'])}` |")
    lines.append(f"| Total size | {evidence['total_size_human']} |")
    lines.append(f"| Total files | {evidence['file_count']} |")
    lines.append(f"| Earliest created seen | {evidence['earliest_created'] or 'Unknown'} |")
    lines.append(f"| Latest modified seen | {evidence['latest_modified'] or 'Unknown'} |")
    lines.append(f"| Eligible files considered | {evidence['eligible_file_estimate']} |")
    lines.append(f"| Evidence files used | {evidence['selected_evidence_count']} |")
    lines.append("")

    lines.append("## Type Breakdown")
    lines.append("")
    lines.append("| Category | Files | Total Size |")
    lines.append("|---|---:|---:|")
    for item in evidence["category_breakdown"]:
        lines.append(f"| {markdown_escape(item['category'])} | {item['files']} | {item['size']} |")
    lines.append("")

    if evidence["direct_child_folders"]:
        lines.append("## Largest Direct Child Folders")
        lines.append("")
        lines.append("| Folder | Files | Size |")
        lines.append("|---|---:|---:|")
        for item in evidence["direct_child_folders"]:
            lines.append(f"| `{markdown_escape(item['path'])}` | {item['files']} | {item['size_human']} |")
        lines.append("")

    lines.append("## Largest Files")
    lines.append("")
    lines.append("| File | Size |")
    lines.append("|---|---:|")
    for item in evidence["largest_files"][:12]:
        lines.append(f"| `{markdown_escape(item['path'])}` | {item['size']} |")
    lines.append("")

    if evidence["signal_files"]:
        lines.append("## High-Signal Files")
        lines.append("")
        for path in evidence["signal_files"][:15]:
            lines.append(f"- `{markdown_escape(path)}`")
        lines.append("")

    if evidence["document_notes"]:
        lines.append("## Selected Document / Text Evidence")
        lines.append("")
        for item in evidence["document_notes"][:12]:
            lines.append(f"- `{markdown_escape(item['path'])}` — {item['summary']}")
        lines.append("")

    if evidence["image_notes"]:
        lines.append("## Selected Image Notes")
        lines.append("")
        for item in evidence["image_notes"][:8]:
            lines.append(f"- `{markdown_escape(item['path'])}` — {item['summary']}")
        lines.append("")

    if narrative.get("keep_delete_move"):
        lines.append("## Review Notes")
        lines.append("")
        for item in narrative["keep_delete_move"][:8]:
            lines.append(f"- {item}")
        lines.append("")

    if evidence["cleanup_candidates"]:
        lines.append("## Obvious Cleanup Candidates")
        lines.append("")
        for item in evidence["cleanup_candidates"][:10]:
            lines.append(f"- `{markdown_escape(item['path'])}` — {item['reason']} ({item['size']})")
        lines.append("")

    lines.append("---")
    lines.append("")
    lines.append("Generated locally by `app.py`. Exact counts, sizes, and dates are computed by the script; AI is only used to interpret the folder from the evidence package.")
    lines.append("")
    return "\n".join(lines)


def render_root_markdown(
    overview: Dict[str, Any],
    root_stats: FolderStats,
    child_reports: Sequence[Dict[str, Any]],
    loose_root_files: Sequence[FileRecord],
    cleanup_payload: Dict[str, List[Dict[str, str]]],
    narrative: Dict[str, Any],
) -> str:
    lines: List[str] = []
    lines.append(f"# {narrative['title']}")
    lines.append("")
    lines.append(
        " ".join(
            [
                badge(f"{overview['total_files']} files", BADGE_COLORS["files"]),
                badge(f"{overview['total_dirs']} folders", BADGE_COLORS["root"]),
                badge(human_size(overview['total_size_bytes']), BADGE_COLORS["size"]),
                badge(f"updated {root_stats.latest_modified or 'unknown'}", BADGE_COLORS["updated"]),
            ]
        )
    )
    lines.append("")
    if narrative.get("top_tags"):
        lines.append(" ".join(badge(tag, BADGE_COLORS["purpose"]) for tag in narrative["top_tags"][:8]))
        lines.append("")
    lines.append(f"> {narrative['summary']}")
    lines.append("")
    lines.append(f"**Overall purpose:** {narrative['purpose']}")
    lines.append("")
    lines.append(narrative["organization_summary"])
    lines.append("")

    lines.append("## Root Snapshot")
    lines.append("")
    lines.append("| Metric | Value |")
    lines.append("|---|---:|")
    lines.append(f"| Root path | `{markdown_escape(overview['root_path'])}` |")
    lines.append(f"| Total size | {human_size(overview['total_size_bytes'])} |")
    lines.append(f"| Total files | {overview['total_files']} |")
    lines.append(f"| Total folders | {overview['total_dirs']} |")
    lines.append(f"| Top-level folders reported | {overview['top_level_folder_count']} |")
    lines.append(f"| Earliest created seen | {root_stats.earliest_created or 'Unknown'} |")
    lines.append(f"| Latest modified seen | {root_stats.latest_modified or 'Unknown'} |")
    lines.append(f"| Report generated | {overview['generated_at']} |")
    lines.append("")

    lines.append("## Top-Level Folder Rollup")
    lines.append("")
    lines.append("| Folder | Files | Size | Summary |")
    lines.append("|---|---:|---:|---|")
    for item in child_reports:
        lines.append(
            f"| `{markdown_escape(item['folder_path'])}` | {item['file_count']} | {item['size_human']} | {markdown_escape(item['summary'])} |"
        )
    lines.append("")

    if loose_root_files:
        lines.append("## Loose Files in Root")
        lines.append("")
        for rec in sorted(loose_root_files, key=lambda r: r.size_bytes, reverse=True)[:20]:
            lines.append(f"- `{markdown_escape(rec.rel_path)}` — {human_size(rec.size_bytes)}")
        lines.append("")

    if narrative.get("delete_recommendations"):
        lines.append("## Recommendations: Review / Delete")
        lines.append("")
        for item in narrative["delete_recommendations"][:12]:
            lines.append(f"- {item}")
        lines.append("")

    if narrative.get("move_recommendations"):
        lines.append("## Recommendations: Move / Reorganize")
        lines.append("")
        for item in narrative["move_recommendations"][:12]:
            lines.append(f"- {item}")
        lines.append("")

    if narrative.get("priority_actions"):
        lines.append("## Priority Actions")
        lines.append("")
        for item in narrative["priority_actions"][:10]:
            lines.append(f"- {item}")
        lines.append("")

    for label, items in cleanup_payload.items():
        if not items:
            continue
        pretty_label = label.replace("_", " ").title()
        lines.append(f"## {pretty_label}")
        lines.append("")
        for item in items[:15]:
            if "path" in item:
                reason = item.get("reason", "")
                size = item.get("size", "")
                suffix = f" — {reason}" if reason else ""
                if size:
                    suffix += f" ({size})"
                lines.append(f"- `{markdown_escape(item['path'])}`{suffix}")
            else:
                lines.append(f"- {item}")
        lines.append("")

    lines.append("---")
    lines.append("")
    lines.append("Generated locally by `app.py`. Per-folder reports were built from recursive scans of each top-level subtree; this root report summarizes those outputs.")
    lines.append("")
    return "\n".join(lines)


def unique_markdown_path(output_dir: Path, base_name: str) -> Path:
    stem = safe_slug(base_name)
    candidate = output_dir / f"{stem}-Markup.md"
    idx = 2
    while candidate.exists():
        candidate = output_dir / f"{stem}-{idx}-Markup.md"
        idx += 1
    return candidate


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create one Markdown report per top-level folder plus a full root report.")
    parser.add_argument("root", help="Root folder to scan")
    parser.add_argument("--output-dir", default="Folder-Markups", help="Single folder where all Markdown files are written")
    parser.add_argument("--include-hidden", action="store_true", help="Include hidden files/folders")
    parser.add_argument("--exclude-output-dir-from-scan", action="store_true", default=True, help="Exclude the output folder if it sits inside the scanned root")
    parser.add_argument("--analyze-images", action="store_true", help="Use the local vision model for selected images")
    parser.add_argument("--disable-ai", action="store_true", help="Skip local AI and write deterministic summaries only")
    parser.add_argument("--ollama-url", default="http://localhost:11434", help="Ollama base URL")
    parser.add_argument("--model", default="gemma3:12b", help="Local model name in Ollama")
    parser.add_argument("--inventory-json", default=None, help="Optional JSON manifest path")
    parser.add_argument("--max-text-file-mb", type=int, default=25, help="Do not extract text from files larger than this many MB")
    parser.add_argument("--max-excerpt-chars", type=int, default=12000, help="Max chars extracted per text-heavy file")
    parser.add_argument("--folder-analysis-ratio", type=float, default=0.65, help="Target fraction of eligible files sampled as evidence for each top-level folder")
    parser.add_argument("--max-evidence-files-per-folder", type=int, default=0, help="Optional cap on evidence files per top-level folder. Use 0 for no cap.")
    parser.add_argument("--max-doc-evidence-per-folder", type=int, default=16, help="How many extracted text snippets to include per folder")
    parser.add_argument("--max-image-evidence-per-folder", type=int, default=8, help="How many selected images to describe per folder")
    parser.add_argument("--max-representative-files", type=int, default=20, help="Representative filenames to keep per folder payload")
    parser.add_argument("--max-largest-files", type=int, default=15, help="Largest files to keep per folder")
    parser.add_argument("--max-recent-files", type=int, default=12, help="Recent files to keep per folder")
    parser.add_argument("--ai-timeout", type=int, default=300, help="Timeout per local model request, in seconds")
    parser.add_argument("--num-ctx", "--context", dest="num_ctx", type=int, default=32768, help="Ollama context window to request per API call")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    root = Path(args.root).resolve()
    if not root.exists() or not root.is_dir():
        print(f"Root folder does not exist or is not a directory: {root}", file=sys.stderr)
        return 1

    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"[1/6] Scanning recursively: {root}")
    scanner = FolderScanner(args)
    overview = scanner.scan()
    print(f"      Found {overview['total_files']} files across {overview['total_dirs']} folders.")
    print(f"      Top-level folders to report: {overview['top_level_folder_count']}")

    ai: Optional[OllamaClient] = None
    if not args.disable_ai:
        try:
            ai = OllamaClient(base_url=args.ollama_url, model=args.model, timeout=args.ai_timeout, num_ctx=args.num_ctx)
            print(f"[2/6] Local AI enabled via Ollama at {args.ollama_url} using model {args.model} (num_ctx={args.num_ctx})")
        except Exception as exc:
            print(f"[2/6] AI initialization failed: {exc}")
            ai = None
    else:
        print("[2/6] Local AI disabled; deterministic mode only.")

    child_reports: List[Dict[str, Any]] = []
    written_files: List[str] = []

    print("[3/6] Building one report per top-level folder")
    for idx, folder_rel in enumerate(scanner.top_level_folders, start=1):
        records = scanner.top_folder_files.get(folder_rel, [])
        if not records:
            continue
        print(f"      [{idx}/{len(scanner.top_level_folders)}] {folder_rel} ({len(records)} files)")
        evidence = build_folder_evidence(folder_rel, records, scanner.folder_stats, args, ai)
        try:
            narrative = generate_folder_narrative(ai, evidence)
        except Exception as exc:
            print(f"         AI folder narrative failed for {folder_rel}: {exc}")
            narrative = generate_folder_narrative(None, evidence)

        md = render_folder_markdown(evidence, narrative)
        folder_output_path = unique_markdown_path(output_dir, Path(folder_rel).name)
        folder_output_path.write_text(md, encoding="utf-8")
        written_files.append(str(folder_output_path))
        child_reports.append({
            "folder_path": folder_rel,
            "label": narrative.get("label") or Path(folder_rel).name,
            "file_count": evidence["file_count"],
            "size_human": evidence["total_size_human"],
            "size_bytes": evidence["total_size_bytes"],
            "summary": narrative["summary"],
            "purpose": narrative["purpose"],
            "top_tags": narrative.get("top_tags", []),
            "report_file": folder_output_path.name,
        })

    child_reports.sort(key=lambda item: item["size_bytes"], reverse=True)

    print("[4/6] Building cleanup signals for the full root report")
    cleanup_payload = build_root_cleanup_candidates(scanner.file_records)

    print("[5/6] Building the full root Markdown")
    loose_root_files = [r for r in scanner.top_folder_files.get(".", [])]
    root_payload = {
        "overview": {
            **overview,
            "total_size_human": human_size(overview["total_size_bytes"]),
        },
        "root_stats": {
            "latest_modified": scanner.folder_stats["."].latest_modified,
            "earliest_created": scanner.folder_stats["."].earliest_created,
        },
        "child_folders": child_reports,
        "loose_root_files": [
            {"path": r.rel_path, "size": human_size(r.size_bytes), "category": r.category}
            for r in sorted(loose_root_files, key=lambda r: r.size_bytes, reverse=True)[:20]
        ],
        "cleanup_candidates": cleanup_payload,
    }
    try:
        root_narrative = generate_root_narrative(ai, root_payload)
    except Exception as exc:
        print(f"      AI root narrative failed: {exc}")
        root_narrative = generate_root_narrative(None, root_payload)

    root_output_path = unique_markdown_path(output_dir, root.name)
    root_md = render_root_markdown(
        overview=overview,
        root_stats=scanner.folder_stats["."],
        child_reports=child_reports,
        loose_root_files=loose_root_files,
        cleanup_payload=cleanup_payload,
        narrative=root_narrative,
    )
    root_output_path.write_text(root_md, encoding="utf-8")
    written_files.append(str(root_output_path))
    print(f"      Wrote full root report to: {root_output_path}")

    print("[6/6] Finalizing optional manifest")
    if args.inventory_json:
        manifest = {
            "overview": {**overview, "total_size_human": human_size(overview["total_size_bytes"])},
            "output_dir": str(output_dir),
            "generated_files": written_files,
            "child_reports": child_reports,
            "cleanup_candidates": cleanup_payload,
        }
        Path(args.inventory_json).resolve().write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"      Wrote manifest JSON to: {Path(args.inventory_json).resolve()}")

    print("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
