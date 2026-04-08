#!/usr/bin/env python3
"""
Generate a polished Markdown dashboard for a folder tree.

What it does:
- Recursively scans a folder
- Collects exact metadata: sizes, timestamps, counts, file types
- Extracts text from PDFs, DOCX, TXT/MD/code/CSV/JSON and some spreadsheet data
- Optionally asks a local multimodal model to summarize documents and images
- Uses the local model again to write human-friendly folder labels / blurbs
- Renders a colorful Markdown report with exact metrics and AI-written prose

Designed for local use on Windows/macOS/Linux.
Default AI backend: Ollama running on http://localhost:11434
"""

from __future__ import annotations
import warnings
warnings.filterwarnings(
    "ignore",
    message=r"Print area cannot be set to Defined name: .*",
    category=UserWarning,
)
import argparse
import base64
import collections
import html
import json
import math
import mimetypes
import os
import re
import sys
import textwrap
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    import requests
except ImportError:  # pragma: no cover
    requests = None

try:
    import fitz  # PyMuPDF
except ImportError:  # pragma: no cover
    fitz = None

try:
    from docx import Document as DocxDocument
except ImportError:  # pragma: no cover
    DocxDocument = None

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

try:
    from PIL import Image
except ImportError:  # pragma: no cover
    Image = None


TEXT_EXTENSIONS = {
    ".txt", ".md", ".markdown", ".rst", ".log", ".ini", ".cfg", ".conf", ".toml",
    ".yaml", ".yml", ".json", ".jsonl", ".csv", ".tsv", ".xml", ".html", ".htm",
    ".css", ".js", ".ts", ".jsx", ".tsx", ".py", ".java", ".c", ".cpp", ".h", ".hpp",
    ".cs", ".go", ".rs", ".rb", ".php", ".sql", ".sh", ".bat", ".ps1", ".tex", ".rtf",
}

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp", ".tif", ".tiff", ".heic"}
VIDEO_EXTENSIONS = {".mp4", ".mkv", ".mov", ".avi", ".wmv", ".m4v", ".webm", ".flv", ".mpeg", ".mpg"}
AUDIO_EXTENSIONS = {".mp3", ".wav", ".flac", ".aac", ".m4a", ".ogg", ".wma", ".aiff"}
ARCHIVE_EXTENSIONS = {".zip", ".rar", ".7z", ".tar", ".gz", ".bz2", ".xz"}
DOCUMENT_EXTENSIONS = {".pdf", ".docx", ".doc", ".pptx", ".ppt", ".xlsx", ".xls", ".epub"}
FONT_EXTENSIONS = {".ttf", ".otf", ".woff", ".woff2"}
EXECUTABLE_EXTENSIONS = {".exe", ".msi", ".app", ".dmg", ".pkg", ".apk", ".ipa"}
CAD_EXTENSIONS = {".blend", ".obj", ".fbx", ".stl", ".step", ".stp", ".dwg", ".dxf"}


BADGE_COLORS = {
    "root": "#2563eb",
    "files": "#0f766e",
    "size": "#7c3aed",
    "docs": "#b45309",
    "images": "#be185d",
    "video": "#b91c1c",
    "audio": "#15803d",
    "code": "#374151",
    "updated": "#0369a1",
    "purpose": "#4f46e5",
}


@dataclass
class FileRecord:
    rel_path: str
    abs_path: str
    name: str
    extension: str
    category: str
    size_bytes: int
    created: Optional[str]
    modified: Optional[str]
    mime_type: str
    is_hidden: bool
    excerpt: Optional[str] = None
    ai_summary: Optional[str] = None
    image_description: Optional[str] = None
    width: Optional[int] = None
    height: Optional[int] = None
    parse_error: Optional[str] = None


@dataclass
class FolderStats:
    rel_path: str
    depth: int
    total_size_bytes: int = 0
    file_count: int = 0
    category_counts: collections.Counter = field(default_factory=collections.Counter)
    extension_counts: collections.Counter = field(default_factory=collections.Counter)
    latest_modified: Optional[str] = None
    earliest_created: Optional[str] = None
    largest_files: List[Tuple[str, int]] = field(default_factory=list)
    sample_files: List[str] = field(default_factory=list)
    doc_summaries: List[Dict[str, str]] = field(default_factory=list)
    image_summaries: List[Dict[str, str]] = field(default_factory=list)


class LocalAIError(RuntimeError):
    pass


class OllamaClient:
    def __init__(self, base_url: str, model: str, timeout: int = 180, keep_alive: str = "30m") -> None:
        if requests is None:
            raise LocalAIError("The 'requests' package is required for AI integration. Install it with: pip install requests")
        self.base_url = base_url.rstrip("/")
        self.model = model
        self.timeout = timeout
        self.keep_alive = keep_alive

    def generate_text(
        self,
        prompt: str,
        system: Optional[str] = None,
        images: Optional[List[Path]] = None,
        schema: Optional[Dict[str, Any]] = None,
        temperature: float = 0.1,
    ) -> str:
        payload: Dict[str, Any] = {
            "model": self.model,
            "prompt": prompt,
            "stream": False,
            "keep_alive": self.keep_alive,
            "options": {"temperature": temperature},
        }
        if system:
            payload["system"] = system
        if schema:
            payload["format"] = schema
        if images:
            payload["images"] = [base64.b64encode(p.read_bytes()).decode("utf-8") for p in images]

        try:
            response = requests.post(f"{self.base_url}/api/generate", json=payload, timeout=self.timeout)
            response.raise_for_status()
        except Exception as exc:  # pragma: no cover
            raise LocalAIError(f"Could not reach local Ollama server at {self.base_url}. {exc}") from exc

        data = response.json()
        text = data.get("response", "")
        if not text:
            raise LocalAIError("The local model returned an empty response.")
        return text.strip()

    def generate_json(
        self,
        prompt: str,
        system: Optional[str] = None,
        images: Optional[List[Path]] = None,
        schema: Optional[Dict[str, Any]] = None,
        temperature: float = 0.1,
    ) -> Any:
        response_text = self.generate_text(
            prompt=prompt,
            system=system,
            images=images,
            schema=schema or {"type": "object"},
            temperature=temperature,
        )
        try:
            return json.loads(response_text)
        except json.JSONDecodeError as exc:
            raise LocalAIError(f"Model did not return valid JSON. First 400 chars: {response_text[:400]!r}") from exc


# -----------------------------
# Utility helpers
# -----------------------------

def human_size(num_bytes: int) -> str:
    if num_bytes < 1024:
        return f"{num_bytes} B"
    units = ["KB", "MB", "GB", "TB", "PB"]
    value = float(num_bytes)
    for unit in units:
        value /= 1024.0
        if value < 1024 or unit == units[-1]:
            return f"{value:.2f} {unit}"
    return f"{num_bytes} B"



def format_dt(ts: Optional[float]) -> Optional[str]:
    if ts is None:
        return None
    try:
        return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return None



def safe_creation_timestamp(stat_result: os.stat_result) -> Optional[float]:
    if hasattr(stat_result, "st_birthtime"):
        return getattr(stat_result, "st_birthtime")
    # On Windows this is creation time. On Linux it is metadata-change time.
    return getattr(stat_result, "st_ctime", None)



def truncate(text: str, max_chars: int) -> str:
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) <= max_chars:
        return text
    return text[: max_chars - 1].rstrip() + "…"



def sanitize_excerpt(text: str, max_chars: int = 2400) -> str:
    text = text.replace("\x00", " ")
    text = re.sub(r"\s+", " ", text)
    return truncate(text, max_chars)



def is_hidden_path(path: Path) -> bool:
    return any(part.startswith(".") for part in path.parts if part not in (".", ".."))



def guess_category(path: Path, mime_type: str) -> str:
    ext = path.suffix.lower()
    if ext in IMAGE_EXTENSIONS or mime_type.startswith("image/"):
        return "image"
    if ext in VIDEO_EXTENSIONS or mime_type.startswith("video/"):
        return "video"
    if ext in AUDIO_EXTENSIONS or mime_type.startswith("audio/"):
        return "audio"
    if ext in ARCHIVE_EXTENSIONS:
        return "archive"
    if ext in DOCUMENT_EXTENSIONS:
        return "document"
    if ext in FONT_EXTENSIONS:
        return "font"
    if ext in EXECUTABLE_EXTENSIONS or mime_type in {"application/x-msdownload", "application/x-dosexec"}:
        return "executable"
    if ext in CAD_EXTENSIONS:
        return "3d/cad"
    if ext in TEXT_EXTENSIONS or mime_type.startswith("text/"):
        if ext in {".py", ".js", ".ts", ".java", ".c", ".cpp", ".h", ".hpp", ".cs", ".go", ".rs", ".rb", ".php", ".sql", ".sh", ".bat", ".ps1", ".css", ".html", ".htm", ".xml", ".json", ".yaml", ".yml", ".toml"}:
            return "code/config"
        return "text/data"
    return "other"



def sample_indices(total: int, wanted: int) -> List[int]:
    if total <= 0:
        return []
    if total <= wanted:
        return list(range(total))
    if wanted == 1:
        return [0]
    step = (total - 1) / (wanted - 1)
    values = sorted({round(i * step) for i in range(wanted)})
    return [min(total - 1, max(0, i)) for i in values]



def safe_read_text_file(path: Path, max_chars: int = 12000) -> str:
    raw = path.read_bytes()[: max_chars * 4]
    for encoding in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
        try:
            return sanitize_excerpt(raw.decode(encoding), max_chars=max_chars)
        except Exception:
            continue
    return ""



def summarize_tabular_preview(rows: List[List[Any]], max_rows: int = 6, max_cols: int = 6) -> str:
    lines: List[str] = []
    for row in rows[:max_rows]:
        clean = [truncate(str(cell), 40) for cell in row[:max_cols] if cell is not None]
        if clean:
            lines.append(" | ".join(clean))
    return sanitize_excerpt(" ; ".join(lines), max_chars=2200)


# -----------------------------
# Document extraction
# -----------------------------

def extract_pdf_excerpt(path: Path, max_chars: int = 12000, max_pages_sampled: int = 10) -> str:
    if fitz is None:
        return ""
    try:
        doc = fitz.open(path)
    except Exception:
        return ""

    text_parts: List[str] = []
    try:
        page_numbers = sample_indices(len(doc), max_pages_sampled)
        for idx in page_numbers:
            try:
                page = doc[idx]
                # sort=True helps restore a more natural reading order.
                text = page.get_text("text", sort=True)
                if text:
                    text_parts.append(text)
            except Exception:
                continue
    finally:
        doc.close()

    return sanitize_excerpt("\n\n".join(text_parts), max_chars=max_chars)



def extract_docx_excerpt(path: Path, max_chars: int = 12000) -> str:
    if DocxDocument is None:
        return ""
    try:
        doc = DocxDocument(str(path))
        text = "\n".join(p.text for p in doc.paragraphs if p.text and p.text.strip())
        return sanitize_excerpt(text, max_chars=max_chars)
    except Exception:
        return ""



def extract_xlsx_excerpt(path: Path, max_chars: int = 12000) -> str:
    if load_workbook is None:
        return ""
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception:
        return ""

    fragments: List[str] = []
    try:
        for ws in list(wb.worksheets)[:4]:
            rows: List[List[Any]] = []
            for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
                rows.append(list(row[:8]))
            preview = summarize_tabular_preview(rows)
            sheet_name = ws.title
            if preview:
                fragments.append(f"Sheet '{sheet_name}': {preview}")
            else:
                fragments.append(f"Sheet '{sheet_name}' appears mostly empty or not text-heavy.")
    finally:
        wb.close()

    return sanitize_excerpt("\n".join(fragments), max_chars=max_chars)



def extract_textual_excerpt(path: Path, max_chars: int = 12000) -> str:
    ext = path.suffix.lower()
    if ext == ".pdf":
        return extract_pdf_excerpt(path, max_chars=max_chars)
    if ext == ".docx":
        return extract_docx_excerpt(path, max_chars=max_chars)
    if ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return extract_xlsx_excerpt(path, max_chars=max_chars)
    if ext in TEXT_EXTENSIONS:
        return safe_read_text_file(path, max_chars=max_chars)
    if fitz is not None and ext in {".epub", ".xps", ".oxps", ".cbz", ".fb2"}:
        return extract_pdf_excerpt(path, max_chars=max_chars)
    return ""



def get_image_dimensions(path: Path) -> Tuple[Optional[int], Optional[int]]:
    if Image is None:
        return None, None
    try:
        with Image.open(path) as img:
            return img.width, img.height
    except Exception:
        return None, None


# -----------------------------
# Core scanner
# -----------------------------

def add_largest_file(bucket: List[Tuple[str, int]], rel_path: str, size_bytes: int, limit: int = 12) -> None:
    bucket.append((rel_path, size_bytes))
    bucket.sort(key=lambda item: item[1], reverse=True)
    del bucket[limit:]



def maybe_add_capped(items: List[Any], value: Any, limit: int) -> None:
    if len(items) < limit:
        items.append(value)


class FolderScanner:
    def __init__(self, args: argparse.Namespace) -> None:
        self.args = args
        self.root = Path(args.root).resolve()
        self.root_name = self.root.name or str(self.root)
        self.folder_stats: Dict[str, FolderStats] = {}
        self.file_records: List[FileRecord] = []
        self.doc_candidates: List[FileRecord] = []
        self.image_candidates: List[FileRecord] = []

    def ensure_folder(self, rel_path: str) -> FolderStats:
        if rel_path not in self.folder_stats:
            depth = 0 if rel_path == "." else rel_path.count(os.sep) + 1
            self.folder_stats[rel_path] = FolderStats(rel_path=rel_path, depth=depth)
        return self.folder_stats[rel_path]

    def iter_paths(self) -> Iterable[Path]:
        for path in self.root.rglob("*"):
            if not self.args.include_hidden and is_hidden_path(path.relative_to(self.root)):
                continue
            if path.is_file():
                yield path

    def ancestor_keys(self, rel_path: Path) -> List[str]:
        keys = ["."]
        parents = rel_path.parents
        # rel file: a/b/c.txt -> want '.', 'a', 'a/b'
        for parent in reversed(parents):
            if str(parent) in {"", "."}:
                continue
            keys.append(str(parent))
        return keys

    def scan(self) -> Dict[str, Any]:
        self.ensure_folder(".")
        total_files = 0
        total_dirs = 0

        # Pre-count dirs for the root tree.
        for dirpath, dirnames, _ in os.walk(self.root):
            rel_dir = os.path.relpath(dirpath, self.root)
            if rel_dir == ".":
                rel_dir = "."
            if not self.args.include_hidden:
                dirnames[:] = [d for d in dirnames if not d.startswith(".")]
            if rel_dir != ".":
                total_dirs += 1
                self.ensure_folder(rel_dir)

        for path in self.iter_paths():
            total_files += 1
            rel_path = path.relative_to(self.root)
            rel_str = str(rel_path)
            ext = path.suffix.lower()
            mime_type = mimetypes.guess_type(str(path))[0] or "application/octet-stream"
            category = guess_category(path, mime_type)
            hidden = is_hidden_path(rel_path)

            try:
                st = path.stat()
            except OSError:
                continue

            created = format_dt(safe_creation_timestamp(st))
            modified = format_dt(st.st_mtime)
            size_bytes = st.st_size

            record = FileRecord(
                rel_path=rel_str,
                abs_path=str(path),
                name=path.name,
                extension=ext or "(none)",
                category=category,
                size_bytes=size_bytes,
                created=created,
                modified=modified,
                mime_type=mime_type,
                is_hidden=hidden,
            )

            if category == "image":
                record.width, record.height = get_image_dimensions(path)
                if self.args.analyze_images:
                    self.image_candidates.append(record)

            should_extract = False
            if category in {"document", "text/data", "code/config"}:
                if size_bytes <= self.args.max_text_file_mb * 1024 * 1024:
                    should_extract = True

            if should_extract:
                try:
                    record.excerpt = extract_textual_excerpt(path, max_chars=self.args.max_excerpt_chars)
                except Exception as exc:
                    record.parse_error = f"Extraction failed: {exc}"
                if record.excerpt:
                    self.doc_candidates.append(record)

            self.file_records.append(record)

            for key in self.ancestor_keys(rel_path):
                fs = self.ensure_folder(key)
                fs.file_count += 1
                fs.total_size_bytes += size_bytes
                fs.category_counts[category] += 1
                fs.extension_counts[ext or "(none)"] += 1

                add_largest_file(fs.largest_files, rel_str, size_bytes, limit=self.args.max_largest_files)
                maybe_add_capped(fs.sample_files, rel_str, self.args.max_sample_files_per_folder)

                if modified and (fs.latest_modified is None or modified > fs.latest_modified):
                    fs.latest_modified = modified
                if created and (fs.earliest_created is None or created < fs.earliest_created):
                    fs.earliest_created = created

        root_stats = self.folder_stats["."]
        overview = {
            "root_path": str(self.root),
            "root_name": self.root_name,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "total_files": total_files,
            "total_dirs": total_dirs,
            "total_size_bytes": root_stats.total_size_bytes,
            "folder_count_in_report": len(self.folder_stats),
        }
        return overview


# -----------------------------
# AI enrichment
# -----------------------------

def summarize_document_with_ai(ai: OllamaClient, record: FileRecord) -> Optional[str]:
    if not record.excerpt:
        return None

    system = (
        "You summarize local files for a folder dashboard. "
        "Be factual and conservative. Do not invent. "
        "If the excerpt is unclear, say so briefly."
    )
    schema = {
        "type": "object",
        "properties": {
            "summary": {"type": "string"},
            "tags": {"type": "array", "items": {"type": "string"}},
        },
        "required": ["summary", "tags"],
    }
    prompt = textwrap.dedent(
        f"""
        File path: {record.rel_path}
        Category: {record.category}
        Extension: {record.extension}
        Size: {human_size(record.size_bytes)}

        Excerpt:
        {record.excerpt}

        Return JSON with:
        - summary: one concise factual sentence about what this file appears to contain
        - tags: 2 to 5 short topic tags
        """
    ).strip()

    try:
        data = ai.generate_json(prompt=prompt, system=system, schema=schema, temperature=0.05)
        summary = str(data.get("summary", "")).strip()
        tags = [str(t).strip() for t in data.get("tags", []) if str(t).strip()]
        if tags:
            summary = f"{summary} Tags: {', '.join(tags[:5])}."
        return summary
    except Exception:
        return None



def describe_image_with_ai(ai: OllamaClient, record: FileRecord) -> Optional[str]:
    image_path = Path(record.abs_path)
    if not image_path.exists():
        return None

    system = (
        "You describe images for a local folder dashboard. "
        "Be concise, literal, and non-speculative. "
        "Do not name people unless the identity is obvious from the image itself."
    )
    schema = {
        "type": "object",
        "properties": {
            "description": {"type": "string"},
            "tags": {"type": "array", "items": {"type": "string"}},
        },
        "required": ["description", "tags"],
    }
    prompt = textwrap.dedent(
        f"""
        Describe this image for a folder inventory.

        File path: {record.rel_path}
        File name: {record.name}
        Dimensions: {record.width or '?'}x{record.height or '?'}

        Return JSON with:
        - description: one short factual sentence
        - tags: 2 to 5 short labels
        """
    ).strip()

    try:
        data = ai.generate_json(prompt=prompt, system=system, images=[image_path], schema=schema, temperature=0.05)
        desc = str(data.get("description", "")).strip()
        tags = [str(t).strip() for t in data.get("tags", []) if str(t).strip()]
        if tags:
            desc = f"{desc} Tags: {', '.join(tags[:5])}."
        return desc
    except Exception:
        return None



def build_ai_payload(
    overview: Dict[str, Any],
    file_records: List[FileRecord],
    folder_stats: Dict[str, FolderStats],
    max_folders: int,
    max_docs_per_folder: int,
    max_images_per_folder: int,
) -> Dict[str, Any]:
    folders: List[Dict[str, Any]] = []
    sorted_folders = sorted(
        folder_stats.values(),
        key=lambda fs: (fs.depth, -fs.total_size_bytes, fs.rel_path),
    )

    for fs in sorted_folders[:max_folders]:
        folders.append(
            {
                "path": fs.rel_path,
                "depth": fs.depth,
                "total_size": human_size(fs.total_size_bytes),
                "file_count": fs.file_count,
                "categories": dict(fs.category_counts.most_common()),
                "top_extensions": dict(fs.extension_counts.most_common(8)),
                "sample_files": fs.sample_files[:12],
                "doc_summaries": fs.doc_summaries[:max_docs_per_folder],
                "image_summaries": fs.image_summaries[:max_images_per_folder],
                "largest_files": [
                    {"path": path, "size": human_size(size)}
                    for path, size in fs.largest_files[:6]
                ],
            }
        )

    top_recent = sorted(
        [f for f in file_records if f.modified],
        key=lambda f: f.modified or "",
        reverse=True,
    )[:15]

    top_largest = sorted(file_records, key=lambda f: f.size_bytes, reverse=True)[:15]

    return {
        "overview": {
            **overview,
            "total_size_human": human_size(int(overview["total_size_bytes"])),
        },
        "largest_files": [
            {"path": f.rel_path, "size": human_size(f.size_bytes), "category": f.category}
            for f in top_largest
        ],
        "recent_files": [
            {"path": f.rel_path, "modified": f.modified, "category": f.category}
            for f in top_recent
        ],
        "folders": folders,
    }



def generate_dashboard_narrative(ai: OllamaClient, payload: Dict[str, Any]) -> Dict[str, Any]:
    system = (
        "You write polished Markdown dashboard content for a local folder inventory. "
        "Never invent metrics. Use only the facts supplied. "
        "Be concrete and helpful. If a folder looks like a media library, say so. "
        "If uncertain, say appears/seems/likely."
    )
    schema = {
        "type": "object",
        "properties": {
            "title": {"type": "string"},
            "overall_summary": {"type": "string"},
            "likely_purpose": {"type": "string"},
            "top_tags": {"type": "array", "items": {"type": "string"}},
            "folders": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "label": {"type": "string"},
                        "summary": {"type": "string"},
                        "what_is_here": {"type": "string"},
                        "tags": {"type": "array", "items": {"type": "string"}},
                    },
                    "required": ["path", "label", "summary", "what_is_here", "tags"],
                },
            },
        },
        "required": ["title", "overall_summary", "likely_purpose", "top_tags", "folders"],
    }

    prompt = textwrap.dedent(
        f"""
        Below is a JSON inventory of a folder tree.

        Your job:
        1. Write a title.
        2. Write a 2-4 sentence overall summary.
        3. Infer the likely purpose of the root folder in one sentence.
        4. Provide 3-8 short tags.
        5. For each folder in the input, write:
           - label: a human-friendly label
           - summary: one concise sentence
           - what_is_here: one slightly richer sentence explaining what seems to live there
           - tags: 2-5 short tags

        Rules:
        - Use only the supplied facts.
        - Do not alter metrics.
        - Be honest about uncertainty.
        - If a folder is mostly media, mention that.
        - If a folder looks like code, backups, papers, spreadsheets, downloads, assets, or personal photos, say so when supported.

        Inventory JSON:
        {json.dumps(payload, ensure_ascii=False)}
        """
    ).strip()

    return ai.generate_json(prompt=prompt, system=system, schema=schema, temperature=0.15)


# -----------------------------
# Markdown rendering
# -----------------------------

def badge(text: str, color: str) -> str:
    escaped = html.escape(text)
    return (
        f"<span style=\"display:inline-block;padding:0.22em 0.58em;margin:0.1em 0.18em 0.1em 0;"
        f"background:{color};color:#fff;border-radius:999px;font-size:0.9em;font-weight:600;\">{escaped}</span>"
    )



def markdown_escape(text: str) -> str:
    return text.replace("|", "\\|")



def build_folder_narrative_lookup(narrative: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    lookup: Dict[str, Dict[str, Any]] = {}
    for item in narrative.get("folders", []):
        path = item.get("path")
        if path:
            lookup[path] = item
    return lookup



def root_type_breakdown(file_records: List[FileRecord]) -> List[Tuple[str, int, int]]:
    counts: Dict[str, Tuple[int, int]] = {}
    for rec in file_records:
        count, size = counts.get(rec.category, (0, 0))
        counts[rec.category] = (count + 1, size + rec.size_bytes)
    rows = [(cat, count, size) for cat, (count, size) in counts.items()]
    rows.sort(key=lambda row: row[2], reverse=True)
    return rows



def render_markdown(
    overview: Dict[str, Any],
    file_records: List[FileRecord],
    folder_stats: Dict[str, FolderStats],
    narrative: Optional[Dict[str, Any]],
    args: argparse.Namespace,
) -> str:
    root = folder_stats["."]
    narrative_lookup = build_folder_narrative_lookup(narrative or {})
    title = (narrative or {}).get("title") or f"Folder Dashboard — {overview['root_name']}"
    overall_summary = (narrative or {}).get("overall_summary") or (
        "This report was generated from a recursive scan of the folder tree. "
        "Counts and sizes below are exact; descriptions are generated conservatively from the scanned contents."
    )
    likely_purpose = (narrative or {}).get("likely_purpose") or "Likely purpose could not be inferred confidently."
    top_tags = (narrative or {}).get("top_tags") or []

    lines: List[str] = []
    lines.append(f"# {title}")
    lines.append("")
    lines.append(
        " ".join(
            [
                badge(f"{overview['total_files']} files", BADGE_COLORS["files"]),
                badge(f"{overview['total_dirs']} folders", BADGE_COLORS["root"]),
                badge(human_size(overview["total_size_bytes"]), BADGE_COLORS["size"]),
                badge(f"updated {root.latest_modified or 'unknown'}", BADGE_COLORS["updated"]),
            ]
        )
    )
    lines.append("")
    if top_tags:
        lines.append(" ".join(badge(tag, BADGE_COLORS["purpose"]) for tag in top_tags[:8]))
        lines.append("")

    lines.append(f"> {overall_summary}")
    lines.append("")
    lines.append(f"**Likely purpose:** {likely_purpose}")
    lines.append("")

    lines.append("## Snapshot")
    lines.append("")
    lines.append("| Metric | Value |")
    lines.append("|---|---:|")
    lines.append(f"| Root path | `{markdown_escape(overview['root_path'])}` |")
    lines.append(f"| Total size | {human_size(overview['total_size_bytes'])} |")
    lines.append(f"| Total files | {overview['total_files']} |")
    lines.append(f"| Total folders | {overview['total_dirs']} |")
    lines.append(f"| Earliest created seen | {root.earliest_created or 'Unknown'} |")
    lines.append(f"| Latest modified seen | {root.latest_modified or 'Unknown'} |")
    lines.append(f"| Report generated | {overview['generated_at']} |")
    lines.append("")

    lines.append("## Type Breakdown")
    lines.append("")
    lines.append("| Category | Files | Total Size |")
    lines.append("|---|---:|---:|")
    for category, count, size in root_type_breakdown(file_records):
        lines.append(f"| {markdown_escape(category)} | {count} | {human_size(size)} |")
    lines.append("")

    lines.append("## Largest Files")
    lines.append("")
    lines.append("| File | Category | Size | Modified |")
    lines.append("|---|---|---:|---|")
    for rec in sorted(file_records, key=lambda f: f.size_bytes, reverse=True)[: args.max_largest_files]:
        lines.append(
            f"| `{markdown_escape(rec.rel_path)}` | {markdown_escape(rec.category)} | {human_size(rec.size_bytes)} | {rec.modified or 'Unknown'} |"
        )
    lines.append("")

    lines.append("## Recent Changes")
    lines.append("")
    lines.append("| File | Category | Modified |")
    lines.append("|---|---|---|")
    for rec in sorted([r for r in file_records if r.modified], key=lambda f: f.modified or "", reverse=True)[:15]:
        lines.append(f"| `{markdown_escape(rec.rel_path)}` | {markdown_escape(rec.category)} | {rec.modified} |")
    lines.append("")

    lines.append("## Folder Sections")
    lines.append("")

    renderable_folders = [
        fs for fs in folder_stats.values()
        if fs.depth <= args.max_section_depth
    ]
    renderable_folders.sort(key=lambda fs: (fs.depth, -fs.total_size_bytes, fs.rel_path))

    for fs in renderable_folders:
        rel = fs.rel_path
        narrative_item = narrative_lookup.get(rel, {})
        label = narrative_item.get("label") or (overview["root_name"] if rel == "." else Path(rel).name)
        summary = narrative_item.get("summary") or ""
        what_is_here = narrative_item.get("what_is_here") or ""
        tags = narrative_item.get("tags") or []

        heading_level = min(3 + fs.depth, 6)
        heading_marks = "#" * heading_level
        display_path = overview["root_name"] if rel == "." else rel
        lines.append(f"{heading_marks} {label}")
        lines.append("")
        lines.append(f"`{markdown_escape(display_path)}`")
        lines.append("")
        folder_badges = [
            badge(f"{fs.file_count} files", BADGE_COLORS["files"]),
            badge(human_size(fs.total_size_bytes), BADGE_COLORS["size"]),
        ]
        for cat, count in fs.category_counts.most_common(4):
            folder_badges.append(badge(f"{cat}: {count}", BADGE_COLORS.get(cat, "#4b5563")))
        lines.append(" ".join(folder_badges))
        lines.append("")
        if tags:
            lines.append(" ".join(badge(tag, BADGE_COLORS["purpose"]) for tag in tags[:5]))
            lines.append("")
        if summary:
            lines.append(f"> {summary}")
            lines.append("")
        if what_is_here:
            lines.append(what_is_here)
            lines.append("")

        lines.append("| Metric | Value |")
        lines.append("|---|---:|")
        lines.append(f"| Files | {fs.file_count} |")
        lines.append(f"| Total size | {human_size(fs.total_size_bytes)} |")
        lines.append(f"| Latest modified | {fs.latest_modified or 'Unknown'} |")
        lines.append(f"| Earliest created | {fs.earliest_created or 'Unknown'} |")
        lines.append("")

        lines.append("**Top extensions:**")
        lines.append("")
        ext_bits = [f"`{markdown_escape(ext)}` × {count}" for ext, count in fs.extension_counts.most_common(10)]
        lines.append(", ".join(ext_bits) if ext_bits else "None")
        lines.append("")

        if fs.doc_summaries:
            lines.append("**Document / text summaries:**")
            lines.append("")
            for item in fs.doc_summaries[: args.max_doc_summaries_per_folder]:
                lines.append(f"- `{markdown_escape(item['path'])}` — {item['summary']}")
            lines.append("")

        if fs.image_summaries:
            lines.append("**Image notes:**")
            lines.append("")
            for item in fs.image_summaries[: args.max_image_summaries_per_folder]:
                lines.append(f"- `{markdown_escape(item['path'])}` — {item['summary']}")
            lines.append("")

        lines.append("<details>")
        lines.append(f"<summary>Show sampled files for <code>{html.escape(display_path)}</code></summary>")
        lines.append("")
        for rel_path in fs.sample_files[: args.max_sample_files_per_folder]:
            lines.append(f"- `{markdown_escape(rel_path)}`")
        lines.append("")
        lines.append("</details>")
        lines.append("")

    lines.append("---")
    lines.append("")
    lines.append("Generated locally by `folder_md_ai_dashboard.py`. Exact metrics are computed by the script; AI text is used only for summaries, labels, and descriptions.")
    lines.append("")
    return "\n".join(lines)


# -----------------------------
# Main program
# -----------------------------

def enrich_folder_stats_with_record(folder_stats: Dict[str, FolderStats], record: FileRecord, args: argparse.Namespace) -> None:
    rel_path = Path(record.rel_path)
    ancestors = ["."]
    for parent in reversed(rel_path.parents):
        if str(parent) in {"", "."}:
            continue
        ancestors.append(str(parent))

    summary_text = record.ai_summary or record.excerpt or record.image_description
    if not summary_text:
        return

    for key in ancestors:
        fs = folder_stats.get(key)
        if not fs:
            continue
        if record.category == "image" and record.image_description:
            maybe_add_capped(
                fs.image_summaries,
                {"path": record.rel_path, "summary": truncate(record.image_description, 280)},
                args.max_image_summaries_per_folder,
            )
        elif record.category in {"document", "text/data", "code/config"} and (record.ai_summary or record.excerpt):
            maybe_add_capped(
                fs.doc_summaries,
                {"path": record.rel_path, "summary": truncate(record.ai_summary or record.excerpt or "", 320)},
                args.max_doc_summaries_per_folder,
            )



def build_fallback_narrative(overview: Dict[str, Any], folder_stats: Dict[str, FolderStats]) -> Dict[str, Any]:
    folders = []
    for fs in sorted(folder_stats.values(), key=lambda item: (item.depth, -item.total_size_bytes, item.rel_path)):
        top_cat = fs.category_counts.most_common(2)
        if top_cat:
            dominant = ", ".join(f"{cat} ({count})" for cat, count in top_cat)
            summary = f"Contains {fs.file_count} files, mainly {dominant}."
        else:
            summary = f"Contains {fs.file_count} files."
        folders.append(
            {
                "path": fs.rel_path,
                "label": overview["root_name"] if fs.rel_path == "." else Path(fs.rel_path).name,
                "summary": summary,
                "what_is_here": "This section is rendered without AI narrative because local AI was disabled or unavailable.",
                "tags": [cat for cat, _ in fs.category_counts.most_common(3)],
            }
        )
    return {
        "title": f"Folder Dashboard — {overview['root_name']}",
        "overall_summary": (
            "This report was generated from a recursive scan of the folder tree. "
            "The local AI summary step was skipped or failed, so the narrative below is deterministic."
        ),
        "likely_purpose": "Purpose not inferred because AI enrichment was unavailable.",
        "top_tags": [cat for cat, _ in folder_stats["."].category_counts.most_common(5)],
        "folders": folders,
    }



def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create a local Markdown dashboard for a folder tree with optional local AI summaries.")
    parser.add_argument("root", help="Root folder to scan")
    parser.add_argument("-o", "--output", default="folder_dashboard.md", help="Output Markdown file")
    parser.add_argument("--inventory-json", default=None, help="Optional JSON dump of the scanned inventory")
    parser.add_argument("--include-hidden", action="store_true", help="Include dotfiles / hidden paths")
    parser.add_argument("--analyze-images", action="store_true", help="Ask the local model to describe images")
    parser.add_argument("--disable-ai", action="store_true", help="Skip local AI and build a deterministic report only")
    parser.add_argument("--ollama-url", default="http://localhost:11434", help="Ollama base URL")
    parser.add_argument("--model", default="gemma3:4b", help="Local model name in Ollama")
    parser.add_argument("--max-text-file-mb", type=int, default=30, help="Do not extract text from files larger than this many MB")
    parser.add_argument("--max-excerpt-chars", type=int, default=12000, help="Max chars extracted per text-heavy file")
    parser.add_argument("--max-docs-to-summarize", type=int, default=36, help="Max number of documents/text files to summarize with AI")
    parser.add_argument("--max-images-to-describe", type=int, default=20, help="Max number of images to describe with AI")
    parser.add_argument("--max-section-depth", type=int, default=2, help="How deep to render folder sections in the Markdown report")
    parser.add_argument("--max-sample-files-per-folder", type=int, default=25, help="Max sampled files shown per folder section")
    parser.add_argument("--max-doc-summaries-per-folder", type=int, default=8, help="Max doc summaries shown per folder section")
    parser.add_argument("--max-image-summaries-per-folder", type=int, default=6, help="Max image summaries shown per folder section")
    parser.add_argument("--max-largest-files", type=int, default=15, help="How many largest files to show")
    parser.add_argument("--ai-timeout", type=int, default=240, help="Timeout per local model request, in seconds")
    return parser.parse_args()



def main() -> int:
    args = parse_args()
    root = Path(args.root)
    if not root.exists() or not root.is_dir():
        print(f"Root folder does not exist or is not a directory: {root}", file=sys.stderr)
        return 1

    scanner = FolderScanner(args)
    print(f"[1/5] Scanning: {root}")
    overview = scanner.scan()
    print(f"      Found {overview['total_files']} files across {overview['total_dirs']} folders.")

    ai: Optional[OllamaClient] = None
    narrative: Optional[Dict[str, Any]] = None

    if not args.disable_ai:
        try:
            ai = OllamaClient(base_url=args.ollama_url, model=args.model, timeout=args.ai_timeout)
            print(f"[2/5] Local AI enabled via Ollama at {args.ollama_url} using model {args.model}")
        except Exception as exc:
            print(f"      AI initialization failed: {exc}")
            ai = None
    else:
        print("[2/5] Local AI disabled; deterministic mode only.")

    if ai:
        print("[3/5] Summarizing text-heavy files with local AI")
        docs = sorted(
            scanner.doc_candidates,
            key=lambda r: (r.size_bytes, len(r.excerpt or "")),
            reverse=True,
        )[: args.max_docs_to_summarize]
        for idx, record in enumerate(docs, start=1):
            print(f"      [{idx}/{len(docs)}] {record.rel_path}")
            summary = summarize_document_with_ai(ai, record)
            if summary:
                record.ai_summary = summary

        if args.analyze_images:
            print("[4/5] Describing images with local vision model")
            imgs = sorted(scanner.image_candidates, key=lambda r: r.size_bytes, reverse=True)[: args.max_images_to_describe]
            for idx, record in enumerate(imgs, start=1):
                print(f"      [{idx}/{len(imgs)}] {record.rel_path}")
                desc = describe_image_with_ai(ai, record)
                if desc:
                    record.image_description = desc
        else:
            print("[4/5] Image analysis skipped")
    else:
        print("[3/5] AI summaries skipped")
        print("[4/5] Image analysis skipped")

    # Push summaries back into folder buckets.
    for record in scanner.file_records:
        enrich_folder_stats_with_record(scanner.folder_stats, record, args)

    print("[5/5] Building Markdown report")
    if ai:
        try:
            payload = build_ai_payload(
                overview=overview,
                file_records=scanner.file_records,
                folder_stats=scanner.folder_stats,
                max_folders=80,
                max_docs_per_folder=args.max_doc_summaries_per_folder,
                max_images_per_folder=args.max_image_summaries_per_folder,
            )
            narrative = generate_dashboard_narrative(ai, payload)
        except Exception as exc:
            print(f"      AI dashboard narrative failed: {exc}")
            narrative = build_fallback_narrative(overview, scanner.folder_stats)
    else:
        narrative = build_fallback_narrative(overview, scanner.folder_stats)

    markdown = render_markdown(
        overview=overview,
        file_records=scanner.file_records,
        folder_stats=scanner.folder_stats,
        narrative=narrative,
        args=args,
    )

    output_path = Path(args.output).resolve()
    output_path.write_text(markdown, encoding="utf-8")
    print(f"      Wrote Markdown report to: {output_path}")

    if args.inventory_json:
        inventory_path = Path(args.inventory_json).resolve()
        serializable_files = [
            {
                "path": r.rel_path,
                "extension": r.extension,
                "category": r.category,
                "size_bytes": r.size_bytes,
                "size_human": human_size(r.size_bytes),
                "created": r.created,
                "modified": r.modified,
                "excerpt": r.excerpt,
                "ai_summary": r.ai_summary,
                "image_description": r.image_description,
                "width": r.width,
                "height": r.height,
            }
            for r in scanner.file_records
        ]
        serializable_folders = {
            rel: {
                "depth": fs.depth,
                "total_size_bytes": fs.total_size_bytes,
                "total_size_human": human_size(fs.total_size_bytes),
                "file_count": fs.file_count,
                "category_counts": dict(fs.category_counts),
                "extension_counts": dict(fs.extension_counts),
                "latest_modified": fs.latest_modified,
                "earliest_created": fs.earliest_created,
                "sample_files": fs.sample_files,
                "doc_summaries": fs.doc_summaries,
                "image_summaries": fs.image_summaries,
                "largest_files": [{"path": p, "size_bytes": s, "size_human": human_size(s)} for p, s in fs.largest_files],
            }
            for rel, fs in scanner.folder_stats.items()
        }
        inventory = {
            "overview": {
                **overview,
                "total_size_human": human_size(overview["total_size_bytes"]),
            },
            "files": serializable_files,
            "folders": serializable_folders,
            "narrative": narrative,
        }
        inventory_path.write_text(json.dumps(inventory, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"      Wrote JSON inventory to: {inventory_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
